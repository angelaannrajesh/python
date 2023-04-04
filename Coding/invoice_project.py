from openpyxl import *
from tkinter import *
wb = load_workbook("C:\\invoice1.xlsx") 
sheet = wb.active
def Excel():
    sheet.cell(row = 1, column = 1).value="My Data"
def focus1(event):
    date_entry_box.focus_set()
def focus2(event):
    customer_name_entry_box.focus_set()
def focus3(event):
    designation_entry_box.focus_set()
def focus4(event):
    p_name1_entry_box.focus_set()
def focus5(event):
    p_name2_entry_box.focus_set()
def focus6(event):
    p_name3_entry_box.focus_set()
def focus7(event):
    p_name4_entry_box.focus_set()
def focus8(event):
    p_name5_entry_box.focus_set()
def focus8(event):
    serial_num_entry_box.focus_set()
def focus9(event):
    qty_entry_box.focus_set()
def focus10(event):
    qty2_entry_box.focus_set()
def focus11(event):
    qty3_entry_box.focus_set()
def focus12(event):
    qty4_entry_box.focus_set()
def focus13(event):
    qty5_entry_box.focus_set()
def focus14(event):
    rate_entry_box.focus_set()
def focus15(event):
    rate1_entry_box.focus_set()
def focus16(event):
    rate2_entry_box.focus_set()
def focus17(event):
    rate3_entry_box.focus_set()
def focus18(event):
    rate4_entry_box.focus_set()
def clear():
    date_entry_box.delete(0,END)
    customer_name_entry_box.delete(0,END)
    designation_entry_box.delete(0,END)
    p_name1_entry_box.delete(0,END)
    p_name2_entry_box.delete(0,END)
    p_name3_entry_box.delete(0,END)
    p_name4_entry_box.delete(0,END)
    p_name5_entry_box.delete(0,END)
    serial_num_entry_box.delete(0,END)
    qty_entry_box.delete(0,END)
    qty2_entry_box.delete(0,END)
    qty3_entry_box.delete(0,END)
    qty4_entry_box.delete(0,END)
    qty5_entry_box.delete(0,END)
    rate_entry_box.delete(0,END)
    rate1_entry_box.delete(0,END)
    rate2_entry_box.delete(0,END)
    rate3_entry_box.delete(0,END)
    rate4_entry_box.delete(0,END)
    
def insert():
    if (date_entry_box.get() == "" and
        customer_name_entry_box.get() == "" and
        designation_entry_box.get() == "" and
        p_name1_entry_box.get()==""and
        p_name2_entry_box.get()==""and
        p_name3_entry_box.get() == "" and
        p_name4_entry_box.get()==""and
        p_name5_entry_box.get()==""and
        serial_num_entry_box.get()==""):
        print("Empty Input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row = 1, column = 1).value = date_entry_box.get()
        sheet.cell(row = 2, column = 1).value = customer_name_entry_box.get()
        sheet.cell(row = 3, column = 1).value = designation_entry_box.get()
        sheet.cell(row = 4, column = 1).value = p_name1_entry_box.get()
        sheet.cell(row = 5, column = 1).value = p_name2_entry_box.get()
        sheet.cell(row = 6, column = 1).value = p_name3_entry_box.get()
        sheet.cell(row = 7, column = 1).value = p_name4_entry_box.get()
        sheet.cell(row = 8, column = 1).value = p_name5_entry_box.get()
        sheet.cell(row = 9, column = 1).value = serial_num_entry_box.get()
        sheet.cell(row = 10, column = 1).value = qty_entry_box.get()
        sheet.cell(row = 11, column = 1).value = qty2_entry_box.get()
        sheet.cell(row = 12, column = 1).value = qty3_entry_box.get()
        sheet.cell(row = 13, column = 1).value = qty4_entry_box.get()
        sheet.cell(row = 14, column = 1).value = qty5_entry_box.get()
        sheet.cell(row = 15, column = 1).value = rate_entry_box.get()
        sheet.cell(row = 16, column = 1).value = rate1_entry_box.get()
        sheet.cell(row = 17, column = 1).value = rate2_entry_box.get()
        sheet.cell(row = 18, column = 1).value = rate3_entry_box.get()
        sheet.cell(row = 19, column = 1).value = rate4_entry_box.get()
        wb.save("C:\\invoice1.xlsx")
        clear()
        
        
if __name__ == "__main__":
    
    
    
  root = Tk()
root.configure(background="cyan")
root.title("mydata")
root.geometry("1000x500")
heading_label = Label(root, text = "Customer Data", bg = "cyan")
heading_label.grid(row = 1, column = 0)
date = Label(root, text = "Date", bg = "cyan")
date.grid(row = 2, column = 0)
date_entry_box = Entry(root)
date_entry_box.grid(row = 2, column = 1,ipadx = "150")
customer_name_label = Label(root, text = "Customer Name", bg = "cyan")
customer_name_label.grid(row = 3, column = 0)
customer_name_entry_box = Entry(root)
customer_name_entry_box.grid(row = 3, column = 1, ipadx = "150")
designation_label = Label(root, text = "Designation", bg = "cyan")
designation_label.grid(row = 4, column = 0)
designation_entry_box = Entry(root)
designation_entry_box.grid(row = 4, column = 1, ipadx = "150")
p_name1_label = Label(root, text = "1. Particular Name", bg = "cyan")
p_name1_label.grid(row = 5, column = 0)
p_name1_entry_box = Entry(root)
p_name1_entry_box.grid(row = 5, column = 1,ipadx = "150")
p_name2_label = Label(root, text = "2. Particular Name", bg = "cyan")
p_name2_label.grid(row = 6, column = 0)
p_name2_entry_box = Entry(root)
p_name2_entry_box.grid(row = 6, column = 1, ipadx = "150")
p_name3_label = Label(root, text = "3. Particular Name", bg = "cyan")
p_name3_label.grid(row = 7, column = 0)
p_name3_entry_box = Entry(root)
p_name3_entry_box.grid(row = 7, column = 1, ipadx = "150")
p_name4_label = Label(root, text = "4. Particular Name", bg = "cyan")
p_name4_label.grid(row = 8, column = 0)
p_name4_entry_box = Entry(root)
p_name4_entry_box.grid(row = 8, column = 1,ipadx = "150")
p_name5_label = Label(root, text = "5. Particular Name", bg ="cyan")
p_name5_label.grid(row = 9, column = 0)
p_name5_entry_box = Entry(root)
p_name5_entry_box.grid(row = 9, column = 1, ipadx = 150)
serial_num_label = Label(root, text = "Serial No" , bg = "cyan")
serial_num_label.grid(row = 2, column = 3)
serial_num_entry_box = Entry(root)
serial_num_entry_box.grid(row = 2, column = 4)
qty_label = Label(root, text = "Qty", bg = "cyan")
qty_label.grid(row = 5, column = 3)
qty_entry_box= Entry(root)
qty_entry_box.grid(row = 5, column = 4)
qty2_label = Label(root, text = "Qty",bg = "cyan")
qty2_label.grid(row = 6, column = 3)
qty2_entry_box= Entry(root)
qty2_entry_box.grid(row = 6, column = 4)
qty3_label = Label(root, text = "Qty",bg = "cyan")
qty3_label.grid(row = 7, column = 3)
qty3_entry_box= Entry(root)
qty3_entry_box.grid(row = 7, column = 4)
qty4_label = Label(root, text = "Qty",bg = "cyan")
qty4_label.grid(row = 8, column = 3)
qty4_entry_box= Entry(root)
qty4_entry_box.grid(row = 8, column = 4)
qty5_label = Label(root, text = "Qty",bg = "cyan")
qty5_label.grid(row = 9, column = 3)
qty5_entry_box= Entry(root)
qty5_entry_box.grid(row = 9, column = 4)
rate_label = Label(root, text = "Rate",bg = "cyan")
rate_label.grid(row = 5, column = 5)
rate_entry_box = Entry(root)
rate_entry_box.grid(row = 5, column = 6)
rate1_label = Label(root, text = "Rate",bg = "cyan")
rate1_label.grid(row = 6, column = 5)
rate1_entry_box = Entry(root)
rate1_entry_box.grid(row = 6, column = 6)
rate2_label = Label(root, text = "Rate",bg = "cyan")
rate2_label.grid(row = 7, column = 5)
rate2_entry_box = Entry(root)
rate2_entry_box.grid(row = 7, column = 6)
rate3_label = Label(root, text = "Rate",bg = "cyan")
rate3_label.grid(row = 8, column = 5)
rate3_entry_box = Entry(root)
rate3_entry_box.grid(row = 8, column = 6)
rate4_label = Label(root, text = "Rate",bg = "cyan")
rate4_label.grid(row = 9, column = 5)
rate4_entry_box = Entry(root)
rate4_entry_box.grid(row = 9, column = 6)

sumbit_button = Button(root, text = 'Sumbit', bg = "light green", command = insert)
sumbit_button.grid(row = 10, column = 1)



Excel()
root.mainloop()
    
